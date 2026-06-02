"""
Представления (views) для веб-интерфейса
"""
import os

from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm

from .models import Category, Supplier, Medication, ConsumableMaterial, Transaction, Batch
from .forms import (
    CategoryForm, SupplierForm, MedicationForm,
    ConsumableMaterialForm, TransactionForm, DateRangeForm
)


def login_view(request):
    """Страница входа"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = AuthenticationForm()
    
    return render(request, 'inventory/login.html', {'form': form})


@login_required
def dashboard(request):
    """Главная панель (дашборд)"""
    # Статистика
    total_medications = Medication.objects.filter(is_active=True).count()
    total_consumables = ConsumableMaterial.objects.filter(is_active=True).count()
    
    # Медикаменты с низким остатком
    low_stock_medications = Medication.objects.filter(
        is_active=True,
        quantity__lte=F('min_quantity')
    ).count()
    
    # Расходники с низким остатком
    low_stock_consumables = ConsumableMaterial.objects.filter(
        is_active=True,
        quantity__lte=F('min_quantity')
    ).count()
    
    # Медикаменты с истекающим сроком годности (30 дней)
    expiring_soon = Medication.objects.filter(
        is_active=True,
        expiry_date__lte=timezone.now().date() + timedelta(days=30),
        expiry_date__gte=timezone.now().date()
    ).count()
    
    # Просроченные медикаменты
    expired = Medication.objects.filter(
        is_active=True,
        expiry_date__lt=timezone.now().date()
    ).count()

    # Расходники с истекающим сроком годности (30 дней)
    expiring_soon_consumables = ConsumableMaterial.objects.filter(
        is_active=True,
        expiry_date__lte=timezone.now().date() + timedelta(days=30),
        expiry_date__gte=timezone.now().date()
    ).count()

    # Просроченные расходники
    expired_consumables = ConsumableMaterial.objects.filter(
        is_active=True,
        expiry_date__lt=timezone.now().date()
    ).count()
    
    # Общая стоимость товаров на складе
    # Стоимость медикаментов считаем по себестоимости партий (точнее статичной цены)
    medications_value = Batch.objects.filter(medication__is_active=True).aggregate(
        total=Sum(F('quantity_remaining') * F('price'))
    )['total'] or 0
    
    consumables_value = ConsumableMaterial.objects.filter(is_active=True).aggregate(
        total=Sum(F('quantity') * F('price'))
    )['total'] or 0
    
    total_inventory_value = medications_value + consumables_value
    
    # Последние операции
    recent_transactions = Transaction.objects.select_related(
        'medication', 'consumable', 'supplier', 'user'
    )[:10]
    
    # Алерты
    alerts = []
    
    if low_stock_medications > 0:
        alerts.append({
            'type': 'warning',
            'icon': '📉',
            'text': f'{low_stock_medications} медикаментов с низким остатком',
            'url': '/medications/?low_stock=1'
        })
    
    if low_stock_consumables > 0:
        alerts.append({
            'type': 'warning',
            'icon': '📉',
            'text': f'{low_stock_consumables} расходников с низким остатком',
            'url': '/consumables/?low_stock=1'
        })
    
    if expiring_soon > 0:
        alerts.append({
            'type': 'info',
            'icon': '⏰',
            'text': f'{expiring_soon} медикаментов истекают в течение 30 дней',
            'url': '/medications/?expiring=1'
        })
    
    if expired > 0:
        alerts.append({
            'type': 'danger',
            'icon': '❌',
            'text': f'{expired} просроченных медикаментов',
            'url': '/medications/?expired=1'
        })

    if expiring_soon_consumables > 0:
        alerts.append({
            'type': 'info',
            'icon': '⏰',
            'text': f'{expiring_soon_consumables} расходников истекают в течение 30 дней',
            'url': '/consumables/?expiring=1'
        })

    if expired_consumables > 0:
        alerts.append({
            'type': 'danger',
            'icon': '❌',
            'text': f'{expired_consumables} просроченных расходников',
            'url': '/consumables/?expired=1'
        })
    
    context = {
        'total_medications': total_medications,
        'total_consumables': total_consumables,
        'low_stock_medications': low_stock_medications,
        'low_stock_consumables': low_stock_consumables,
        'expiring_soon': expiring_soon,
        'expired': expired,
        'total_inventory_value': total_inventory_value,
        'recent_transactions': recent_transactions,
        'alerts': alerts,
    }
    
    return render(request, 'inventory/dashboard.html', context)


@login_required
def medication_list(request):
    """Список медикаментов"""
    medications = Medication.objects.filter(is_active=True).select_related('category')
    
    # Фильтры
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    low_stock = request.GET.get('low_stock', '')
    expiring = request.GET.get('expiring', '')
    expired = request.GET.get('expired', '')
    
    if search:
        medications = medications.filter(
            Q(name__icontains=search) |
            Q(manufacturer__icontains=search) |
            Q(active_ingredient__icontains=search)
        )
    
    if category_id:
        medications = medications.filter(category_id=category_id)
    
    if low_stock:
        medications = medications.filter(quantity__lte=F('min_quantity'))
    
    if expiring:
        medications = medications.filter(
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            expiry_date__gte=timezone.now().date()
        )
    
    if expired:
        medications = medications.filter(expiry_date__lt=timezone.now().date())
    
    categories = Category.objects.all()
    
    context = {
        'medications': medications,
        'categories': categories,
        'search': search,
        'category_id': category_id,
    }
    
    return render(request, 'inventory/medication_list.html', context)


@login_required
def medication_detail(request, pk):
    """Детальная информация о медикаменте"""
    medication = get_object_or_404(Medication, pk=pk)
    transactions = medication.transactions.all()[:20]
    # Партии с фактическим поступлением (в порядке FEFO — ближайший срок первым)
    batches = medication.batches.filter(quantity_received__gt=0)

    context = {
        'medication': medication,
        'transactions': transactions,
        'batches': batches,
    }

    return render(request, 'inventory/medication_detail.html', context)


@login_required
def medication_create(request):
    """Создание медикамента"""
    if request.method == 'POST':
        form = MedicationForm(request.POST)
        if form.is_valid():
            medication = form.save()
            messages.success(request, f'Медикамент "{medication.name}" успешно создан')
            return redirect('medication_detail', pk=medication.pk)
    else:
        form = MedicationForm()
    
    return render(request, 'inventory/medication_form.html', {
        'form': form,
        'title': 'Добавить медикамент'
    })


@login_required
def medication_update(request, pk):
    """Редактирование медикамента"""
    medication = get_object_or_404(Medication, pk=pk)
    
    if request.method == 'POST':
        form = MedicationForm(request.POST, instance=medication)
        if form.is_valid():
            form.save()
            messages.success(request, f'Медикамент "{medication.name}" успешно обновлен')
            return redirect('medication_detail', pk=medication.pk)
    else:
        form = MedicationForm(instance=medication)
    
    return render(request, 'inventory/medication_form.html', {
        'form': form,
        'title': 'Редактировать медикамент',
        'medication': medication
    })


@login_required
def medication_delete(request, pk):
    """Удаление (деактивация) медикамента"""
    medication = get_object_or_404(Medication, pk=pk)
    
    if request.method == 'POST':
        medication.is_active = False
        medication.save()
        messages.success(request, f'Медикамент "{medication.name}" успешно удален')
        return redirect('medication_list')
    
    return render(request, 'inventory/medication_confirm_delete.html', {
        'medication': medication
    })


@login_required
def consumable_list(request):
    """Список расходных материалов"""
    consumables = ConsumableMaterial.objects.filter(is_active=True).select_related('category')
    
    # Фильтры
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    low_stock = request.GET.get('low_stock', '')
    expiring = request.GET.get('expiring', '')
    expired = request.GET.get('expired', '')

    if search:
        consumables = consumables.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )

    if category_id:
        consumables = consumables.filter(category_id=category_id)

    if low_stock:
        consumables = consumables.filter(quantity__lte=F('min_quantity'))

    if expiring:
        consumables = consumables.filter(
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            expiry_date__gte=timezone.now().date()
        )

    if expired:
        consumables = consumables.filter(expiry_date__lt=timezone.now().date())

    categories = Category.objects.all()
    
    context = {
        'consumables': consumables,
        'categories': categories,
        'search': search,
        'category_id': category_id,
    }
    
    return render(request, 'inventory/consumable_list.html', context)


@login_required
def consumable_detail(request, pk):
    """Детальная информация о расходном материале"""
    consumable = get_object_or_404(ConsumableMaterial, pk=pk)
    transactions = consumable.transactions.all()[:20]
    
    context = {
        'consumable': consumable,
        'transactions': transactions,
    }
    
    return render(request, 'inventory/consumable_detail.html', context)


@login_required
def consumable_create(request):
    """Создание расходного материала"""
    if request.method == 'POST':
        form = ConsumableMaterialForm(request.POST)
        if form.is_valid():
            consumable = form.save()
            messages.success(request, f'Материал "{consumable.name}" успешно создан')
            return redirect('consumable_detail', pk=consumable.pk)
    else:
        form = ConsumableMaterialForm()
    
    return render(request, 'inventory/consumable_form.html', {
        'form': form,
        'title': 'Добавить расходный материал'
    })


@login_required
def consumable_update(request, pk):
    """Редактирование расходного материала"""
    consumable = get_object_or_404(ConsumableMaterial, pk=pk)
    
    if request.method == 'POST':
        form = ConsumableMaterialForm(request.POST, instance=consumable)
        if form.is_valid():
            form.save()
            messages.success(request, f'Материал "{consumable.name}" успешно обновлен')
            return redirect('consumable_detail', pk=consumable.pk)
    else:
        form = ConsumableMaterialForm(instance=consumable)
    
    return render(request, 'inventory/consumable_form.html', {
        'form': form,
        'title': 'Редактировать расходный материал',
        'consumable': consumable
    })


@login_required
def consumable_delete(request, pk):
    """Удаление (деактивация) расходного материала"""
    consumable = get_object_or_404(ConsumableMaterial, pk=pk)
    
    if request.method == 'POST':
        consumable.is_active = False
        consumable.save()
        messages.success(request, f'Материал "{consumable.name}" успешно удален')
        return redirect('consumable_list')
    
    return render(request, 'inventory/consumable_confirm_delete.html', {
        'consumable': consumable
    })


@login_required
def transaction_list(request):
    """Список операций"""
    transactions = Transaction.objects.all().select_related(
        'medication', 'consumable', 'supplier', 'user'
    )
    
    # Фильтры
    transaction_type = request.GET.get('type', '')
    item_type = request.GET.get('item_type', '')
    
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    if item_type:
        transactions = transactions.filter(item_type=item_type)
    
    context = {
        'transactions': transactions,
    }
    
    return render(request, 'inventory/transaction_list.html', context)


@login_required
def transaction_create(request):
    """Создание операции"""
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            messages.success(request, 'Операция успешно создана')
            return redirect('transaction_list')
    else:
        form = TransactionForm()
    
    return render(request, 'inventory/transaction_form.html', {
        'form': form,
        'title': 'Новая операция'
    })


@login_required
def transaction_update(request, pk):
    """Редактирование операции (остатки пересчитываются автоматически)"""
    transaction = get_object_or_404(Transaction, pk=pk)

    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            messages.success(request, 'Операция успешно обновлена')
            return redirect('transaction_list')
    else:
        form = TransactionForm(instance=transaction)

    return render(request, 'inventory/transaction_form.html', {
        'form': form,
        'title': 'Редактировать операцию',
        'transaction': transaction,
    })


@login_required
def transaction_delete(request, pk):
    """Удаление операции с откатом её влияния на остаток"""
    transaction = get_object_or_404(Transaction, pk=pk)

    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Операция удалена, остаток на складе скорректирован')
        return redirect('transaction_list')

    return render(request, 'inventory/transaction_confirm_delete.html', {
        'transaction': transaction
    })


# ============================================================
# Категории
# ============================================================

@login_required
def category_list(request):
    """Список категорий"""
    categories = Category.objects.annotate(
        med_count=Count('medications', distinct=True),
        cons_count=Count('consumables', distinct=True),
    ).order_by('name')
    return render(request, 'inventory/category_list.html', {'categories': categories})


@login_required
def category_create(request):
    """Создание категории"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Категория «{category.name}» создана')
            return redirect('category_list')
    else:
        form = CategoryForm()

    return render(request, 'inventory/category_form.html', {
        'form': form,
        'title': 'Добавить категорию',
    })


@login_required
def category_update(request, pk):
    """Редактирование категории"""
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Категория «{category.name}» обновлена')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)

    return render(request, 'inventory/category_form.html', {
        'form': form,
        'title': 'Редактировать категорию',
        'category': category,
    })


@login_required
def category_delete(request, pk):
    """Удаление категории"""
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        name = category.name
        category.delete()
        messages.success(request, f'Категория «{name}» удалена')
        return redirect('category_list')

    context = {
        'category': category,
        'med_count': category.medications.count(),
        'cons_count': category.consumables.count(),
    }
    return render(request, 'inventory/category_confirm_delete.html', context)


# ============================================================
# Поставщики
# ============================================================

@login_required
def supplier_list(request):
    """Список поставщиков"""
    suppliers = Supplier.objects.filter(is_active=True)

    search = request.GET.get('search', '')
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(contact_person__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )

    return render(request, 'inventory/supplier_list.html', {
        'suppliers': suppliers.order_by('name'),
        'search': search,
    })


@login_required
def supplier_create(request):
    """Создание поставщика"""
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'Поставщик «{supplier.name}» создан')
            return redirect('supplier_list')
    else:
        form = SupplierForm()

    return render(request, 'inventory/supplier_form.html', {
        'form': form,
        'title': 'Добавить поставщика',
    })


@login_required
def supplier_update(request, pk):
    """Редактирование поставщика"""
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f'Поставщик «{supplier.name}» обновлён')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'inventory/supplier_form.html', {
        'form': form,
        'title': 'Редактировать поставщика',
        'supplier': supplier,
    })


@login_required
def supplier_delete(request, pk):
    """Удаление (деактивация) поставщика"""
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == 'POST':
        supplier.is_active = False
        supplier.save()
        messages.success(request, f'Поставщик «{supplier.name}» удалён')
        return redirect('supplier_list')

    return render(request, 'inventory/supplier_confirm_delete.html', {
        'supplier': supplier
    })


def _parse_date_range(request):
    """Извлекает диапазон дат из формы фильтра"""
    form = DateRangeForm(request.GET or None)
    date_from = date_to = None
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
    return form, date_from, date_to


def _period_transactions(date_from=None, date_to=None):
    """Операции за период (по дате создания)"""
    txns = Transaction.objects.select_related(
        'medication', 'consumable', 'supplier', 'user'
    )
    if date_from:
        txns = txns.filter(created_at__date__gte=date_from)
    if date_to:
        txns = txns.filter(created_at__date__lte=date_to)
    return txns


def _transaction_period_summary(date_from=None, date_to=None):
    """Сводка по операциям за период"""
    txns = _period_transactions(date_from, date_to)
    amount_expr = ExpressionWrapper(
        F('quantity') * F('price'),
        output_field=DecimalField(max_digits=15, decimal_places=2)
    )

    type_labels = dict(Transaction.TRANSACTION_TYPES)
    by_type = [
        {
            'type': row['transaction_type'],
            'label': type_labels.get(row['transaction_type'], row['transaction_type']),
            'count': row['count'],
            'total_qty': row['total_qty'] or 0,
            'total_amount': row['total_amount'] or 0,
        }
        for row in txns.values('transaction_type').annotate(
            count=Count('id'),
            total_qty=Sum('quantity'),
            total_amount=Sum(amount_expr),
        ).order_by('transaction_type')
    ]

    total_in = txns.filter(transaction_type='in').aggregate(s=Sum(amount_expr))['s'] or 0
    total_out = txns.filter(
        transaction_type__in=['out', 'write_off']
    ).aggregate(s=Sum(amount_expr))['s'] or 0

    return {
        'date_from': date_from,
        'date_to': date_to,
        'count': txns.count(),
        'by_type': by_type,
        'total_in': total_in,
        'total_out': total_out,
        'transactions': txns[:100],
    }


@login_required
def reports(request):
    """Отчеты"""
    form, date_from, date_to = _parse_date_range(request)

    # Общая статистика
    stats = {
        'total_medications': Medication.objects.filter(is_active=True).count(),
        'total_consumables': ConsumableMaterial.objects.filter(is_active=True).count(),
        'total_transactions': Transaction.objects.count(),
        'low_stock_items': (
            Medication.objects.filter(is_active=True, quantity__lte=F('min_quantity')).count() +
            ConsumableMaterial.objects.filter(is_active=True, quantity__lte=F('min_quantity')).count()
        ),
    }

    context = {
        'form': form,
        'stats': stats,
        'period': _transaction_period_summary(date_from, date_to),
        'query_string': request.GET.urlencode(),
    }

    return render(request, 'inventory/reports.html', context)


# Кэш имени зарегистрированного PDF-шрифта в рамках процесса
_PDF_FONT = {}


def _get_pdf_font():
    """Регистрирует и возвращает имя шрифта с поддержкой кириллицы для PDF.

    Ищет подходящий TTF в проекте и в типовых системных каталогах
    (DejaVu — Linux, Arial — macOS/Windows). Если ничего не найдено,
    возвращает Helvetica (без кириллицы) как запасной вариант.
    """
    if 'name' in _PDF_FONT:
        return _PDF_FONT['name']

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        settings.BASE_DIR / 'static' / 'fonts' / 'DejaVuSans.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/TTF/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/Library/Fonts/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
        'C:/Windows/Fonts/arial.ttf',
    ]

    name = 'Helvetica'
    for path in candidates:
        path = str(path)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('PDFCyrillic', path))
                name = 'PDFCyrillic'
                break
            except Exception:
                continue

    _PDF_FONT['name'] = name
    return name


@login_required
def reports_export_excel(request):
    """Экспорт операций за период в Excel (.xlsx)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    _form, date_from, date_to = _parse_date_range(request)
    txns = _period_transactions(date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Операции'

    headers = [
        'Дата', 'Тип операции', 'Тип товара', 'Товар',
        'Количество', 'Цена', 'Сумма', 'Поставщик', 'Пользователь'
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0066CC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for t in txns:
        item = t.medication.name if t.medication else (t.consumable.name if t.consumable else '—')
        ws.append([
            timezone.localtime(t.created_at).strftime('%d.%m.%Y %H:%M'),
            t.get_transaction_type_display(),
            t.get_item_type_display(),
            item,
            float(t.quantity),
            float(t.price),
            float(t.total_amount),
            t.supplier.name if t.supplier else '—',
            t.user.username if t.user else '—',
        ])

    widths = [18, 16, 22, 32, 12, 12, 12, 22, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="aqualab_operations.xlsx"'
    wb.save(response)
    return response


@login_required
def reports_export_pdf(request):
    """Экспорт отчёта по операциям за период в PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    _form, date_from, date_to = _parse_date_range(request)
    summary = _transaction_period_summary(date_from, date_to)
    txns = _period_transactions(date_from, date_to)

    font = _get_pdf_font()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Title'], fontName=font, fontSize=18)
    h2_style = ParagraphStyle('h2', parent=styles['Heading2'], fontName=font)
    normal = ParagraphStyle('n', parent=styles['Normal'], fontName=font, fontSize=10)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="aqualab_report.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm
    )

    elements = [Paragraph('Aqualab — отчёт по операциям', title_style)]

    if date_from or date_to:
        f = date_from.strftime('%d.%m.%Y') if date_from else '…'
        t = date_to.strftime('%d.%m.%Y') if date_to else '…'
        period_txt = f'Период: {f} — {t}'
    else:
        period_txt = 'Период: весь период'
    elements.append(Paragraph(period_txt, normal))
    elements.append(Paragraph(
        f"Всего операций: {summary['count']}. "
        f"Приход на сумму: {summary['total_in']:.2f} c, "
        f"расход и списание: {summary['total_out']:.2f} c.",
        normal
    ))
    elements.append(Spacer(1, 0.5 * cm))

    # Сводка по типам операций
    if summary['by_type']:
        data = [['Тип операции', 'Кол-во', 'Сумма, c']]
        for row in summary['by_type']:
            data.append([row['label'], str(row['count']), f"{row['total_amount']:.2f}"])
        tbl = Table(data, hAlign='LEFT')
        tbl.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef3f9')]),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 0.5 * cm))

    # Журнал операций
    elements.append(Paragraph('Журнал операций', h2_style))
    tdata = [['Дата', 'Тип', 'Товар', 'Кол-во', 'Цена', 'Сумма', 'Пользователь']]
    for t in txns[:500]:
        item = t.medication.name if t.medication else (t.consumable.name if t.consumable else '—')
        tdata.append([
            timezone.localtime(t.created_at).strftime('%d.%m.%Y %H:%M'),
            t.get_transaction_type_display(),
            item,
            f"{t.quantity:g}",
            f"{t.price:.2f}",
            f"{t.total_amount:.2f}",
            t.user.username if t.user else '—',
        ])

    if len(tdata) > 1:
        table = Table(tdata, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#212529')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph('За выбранный период операций не найдено.', normal))

    doc.build(elements)
    return response


@login_required
def about(request):
    """Страница «Об авторе»"""
    return render(request, 'inventory/about.html')
